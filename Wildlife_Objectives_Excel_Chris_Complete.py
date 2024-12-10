import openpyxl
from openpyxl import load_workbook

'''
Define all your variables at the top so you don't have to search through the code to find them and they can easily be changed in one place.
** File paths shouldn't be stored on github for security reasons, we will teach you how to avoid that in the future.
'''
path = r'\\spatialfiles.bcgov\work\srm\nel\Local\Geomatics\Workarea\csostad\GitHub_Repositories\Marina_Cunningham\Script_Folder\FSW_Objectives_Chris_Copy.xlsx'
book = load_workbook(path)
consolidated_sheet = 'All_Data'

# Create a new consolidated sheet at the start of the workbook. 0 is the index of the new sheet using python 0 indexing.
sheet0 = book.create_sheet(consolidated_sheet, 0)

'''
Use lots of print statements to help you debug your code. This will help you understand what is happening and where things are going wrong.
Use f-strings to make your print statements more readable and easier to write.

'''
print(f"New sheet {sheet0} created.")


# Add headers - as your skills develop, you could read the headers from another sheet so you dont need to define them
headers = ["FSW_TAG", "FISH_SENSITIVE_WS_POLY_ID", "OBJECTIVE_1", "OBJECTIVE_2", 
           "OBJECTIVE_3", "OBJECTIVE_4", "OBJECTIVE_5", "OBJECTIVE_6", 
           "OBJECTIVE_7", "OBJECTIVE_8", "OBJECTIVE_9"]


sheet0.append(headers)

print(f"Headers added to {sheet0}.") # Here you can see we use f strings to make our print statements more readable

# Define all the sheets (here we are creating worksheet objects, the sheets are no longer just a string of characters)
sheet1 = book["FSW_ID_1_001_to_F_1_011_edit"]
sheet2 = book["FSW_ID_F_4_001_edit"]
sheet3 = book["FSW_ID_F_6_001_to_F_6_005_edit"]
sheet4 = book["FSW_ID_F_3_001_to_F_3_006_edit"]
sheet5 = book["FSW_ID_F_8_001_to_F_8_008_edit"]
sheet6 = book["FSW_ID_F_7_001_and_F_7_005_edit"]
sheet7 = book["FSW_ID_F_7_002_edit"]
sheet8 = book["FSW_ID_F_7_003_F_7_004_edit"]
sheet9 = book["FSW_ID_F_7_006_to_F_7_008_edit"]
sheet10 = book["FSW_ID_F_7_019_edit"]
sheet11 = book["FSW_ID_F_7_020_to_F_7_023_edit"]
sheet12 = book["FSW_F_3_007_and_F_3_008_edit"]
sheet13 = book["FSW_ID_F_3_009_to_F_3_014_edit"]
sheet14 = book["FSW_ID_F_5_001_edit"]


# Assigning all the sheets to a list that we can iterate over with a for loop
sheets_list = [sheet1, sheet2, sheet3, sheet4, sheet5, sheet6, sheet7, sheet8, 
               sheet9, sheet10, sheet11, sheet12, sheet13, sheet14]

print(f"List of sheets created. Here is the list {sheets_list}")

# Append all data from the sheets to sheet0, skipping the header row

# For each sheet in the list of sheets, assign the sheet to the variable sheet
for each_sheet in sheets_list: 
    
    # iterate over each row in the sheet starting at row 2 and assign the data in the row to the variable one_row
    for one_row in each_sheet.iter_rows(min_row=2, values_only=True): # ...only include values that have data (values_only=True)
        print(f"Appending row {one_row}.")
        
        sheet0.append(one_row) # append that one row of data to the new sheet

# Save the updated workbook
book.save(path)
print("Workbook saved.")

print(f"Data successfully consolidated into {consolidated_sheet} sheet.")
