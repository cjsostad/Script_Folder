import openpyxl
from openpyxl import Workbook, load_workbook

#load an excel sheet
# book = load_workbook(filename = r'\\spatialfiles.bcgov\work\srm\nel\Local\Geomatics\Workarea\csostad\GitHub_Repositories\Marina_Cunningham\Script_Folder\FSW_Objectives__Chris_Copy.xlsx')

'''
To save you from having to keep entering the file path, you can assign it to a variable. 
'''

path = r'\\spatialfiles.bcgov\work\srm\nel\Local\Geomatics\Workarea\csostad\GitHub_Repositories\Marina_Cunningham\Script_Folder\FSW_Objectives__Chris_Copy.xlsx'
book = load_workbook(path)

print(book)


#printsheetnames to verify
print(book.sheetnames)

#View active data (don't quite understand this function, is it needed?) 
'''
This is taking the active sheet open in the excel file and assigning it to the variable all_fsw. By default I think it would be sheet one. 
'''
# all_fsw = book.active

#create new sheets that extract specific data
'''
New woksheet is created with the ALL_FISH_SENS... title and will be placed as the 1st sheet of the workbook (index 0)
'''
sheet0 = book.create_sheet("ALL_FISH_SENSITIVE_WS_POLY_ID", 0)




'''
At this stage, I wanted to make sure the new sheet was being created but you need to save it first before you can see it in the excel file.
'''

book.save(path) # I used the file_path variable I created earlier



#Get the sheetnames in the file (0-12, 14 total)

#Label columns
sheet0["A1"] = "FSW_TAG"
sheet0["B1"] = "FISH_SENSITIVE_WS_POLY_ID"
sheet0["C1"] = "OBJECTIVE_1"
sheet0["D1"] = "OBJECTIVE_2"
sheet0["E1"] = "OBJECTIVE_3"
sheet0["F1"] = "OBJECTIVE_4"
sheet0["G1"] = "OBJECTIVE_5"
sheet0["H1"] = "OBJECTIVE_6"
sheet0["I1"] = "OBJECTIVE_7"
sheet0["J1"] = "OBJECTIVE_8"
sheet0["K1"] = "OBJECTIVE_9"

'''
I like to use print statement everywhere to check every step of the way.  Here I would want it to print the column names from A1 to K1.
I always get super confused when working with python and excel. Python starts counting at 0 and Excel starts counting at 1. So always keep that 
in mind. I will use a for - loop to print the column names from A1 to K1.
'''

# Print column names from A1 to K1
for each_cell in sheet0.iter_cols(min_col=1, max_col=11, min_row=1, max_row=1, values_only=True):
    print(each_cell[0])
'''
The above for loop basically reads, go through each column (.iter_cols) in sheet0, starting from column 1 (min_col=1) to column 11 (max_col=11). Only read column
that have values in them (values_only=True). The min_row and max_row are set to 1 because we only want to read the column names from the first row.(Excel 1 basted indexing)
As the loop goes through each column, it will assign that to the variable "each_cell" and print the first and only value of each_cell. (0 = python 0 based indexing)
'''


# define all the sheets - each of these is creating a python "Object", it is no longer just a string name
sheet1 = book["FSW_ID_1_001_to_F_1_011_edit"]
sheet2 = book["FSW_ID_F_4_001_edit"]
sheet3 = book["FSW_ID_F_6_001_to_F_6_005_edit"]
sheet4 = book["FSW_ID_F_3_001_to_F_3_006_edit"]
sheet5 = book["FSW_ID_F_8_001_to_F_8_008_edit"]
sheet6 = book["FSW_ID_F_7_001_and_F_7_005_edit"]
sheet7 = book["FSW_ID_F_7_002_edit"]
sheet8 = book["FSW_ID_F_7_003_F_7_004_edit"]
sheet9 = book["FSW_ID_F_7_006_to_F_7_008_edit"]
sheet10 = book["FSW_ID_F_7_019_edit"]
sheet11 = book["FSW_ID_F_7_020_to_F_7_023_edit"]
sheet12 = book["FSW_F_3_007_and_F_3_008_edit"]
sheet13 = book["FSW_ID_F_3_009_to_F_3_014_edit"]
sheet14 = book["FSW_ID_F_5_001_edit"]

# Create a list of sheets
sheets_list = [sheet1, sheet2, sheet3, sheet4, sheet5, sheet6, sheet7, sheet8, sheet9, sheet10, sheet11, sheet12, sheet13, sheet14]

#delete first row in all sheets so there is not duplication(keep for later)
'''
We can get away without deleting the first row by having the for loop start at row 2 when readin the data
'''

# sheet1.delete_rows(idx=1, amount=1)
# sheet2.delete_rows(idx=1, amount=1)
# sheet3.delete_rows(idx=1, amount=1)
# sheet4.delete_rows(idx=1, amount=1)
# sheet5.delete_rows(idx=1, amount=1)
# sheet6.delete_rows(idx=1, amount=1)
# sheet7.delete_rows(idx=1, amount=1)
# sheet8.delete_rows(idx=1, amount=1)
# sheet9.delete_rows(idx=1, amount=1)
# sheet10.delete_rows(idx=1, amount=1)
# sheet11.delete_rows(idx=1, amount=1)
# sheet12.delete_rows(idx=1, amount=1)
# sheet13.delete_rows(idx=1, amount=1)
# sheet14.delete_rows(idx=1, amount=1)





'''
Pseudo Code for copying data from sheet1 to sheet0
1. Get the max row count of sheet1
2. Loop through each row in sheet1 starting from row 2 (row 1 is the column names)
3. For each row, get the value of each cell in that row


'''

# Start appending data
for sheet_name in sheets_list:
    sheet = book[sheet_name]
    for row in sheet.iter_rows(values_only=True):
        sheet0.append(row)
#Append the data (didnt work)
#for each_row in sheet1:
    #sheet0.append(each_row)

#book.copy_worksheet(sheet1)
#print a column
#print(sheet1.columns=1)

#Copy over the data from (did not work) sheet1, sheet2, sheet3, sheet4, sheet5, sheet6, sheet7, sheet8, sheet9, sheet10, sheet11, sheet12, sheet13)
#target = book.copy_worksheet(sheet1)
#Append the data from the sheets into all_fsw(this append function didnt work) sheet1, sheet2, sheet3, sheet4, sheet5, sheet6, sheet7, sheet8, sheet9, sheet10, sheet11, sheet12, sheet13)
#all_fsw.append(sheet1) 

#possible code solutions
#Dictionary or classes 
#iterate through all operational sheets to extract row and column data

book.save (filename = r'\\spatialfiles.bcgov\work\srm\nel\Local\Geomatics\Workarea\mcunning\Script_Folder_Marina\FSW_Objectives_Copy_Python.xlsx')



